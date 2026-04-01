#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import os
import glob
from typing import Dict, Tuple, Any, List, Optional, Set

from openpyxl import load_workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.utils import ImageReader

VARIANT_ALIASES = {
    "baseline":     ["baseline"],
    "crops_black":  ["crops_black", "crop_black", "black"],
    "crops_white":  ["crops_white", "crop_white", "white"],
}

def _to_str(x) -> str:
    return "" if x is None else str(x)

def parse_labels(s: Any) -> List[str]:
    if s is None:
        return []
    ss = str(s).strip()
    if ss == "" or ss.lower() == "nan":
        return []
    ss = ss.replace(";", ",")
    parts = [p.strip() for p in ss.split(",") if p.strip()]
    out, seen = [], set()
    for p in parts:
        if p not in seen:
            seen.add(p)
            out.append(p)
    return out

def parse_gt_image_labels(s: Any) -> Set[str]:
    if s is None:
        return set()
    ss = str(s).strip()
    if ss == "" or ss.lower() == "nan":
        return set()
    # GT image-level costuma vir com ';'
    ss = ss.replace(",", ";")
    parts = [p.strip() for p in ss.split(";") if p.strip()]
    return set([p for p in parts])

def compute_status_crop(gt_crop_label: str, pred: List[str]) -> str:
    gt = gt_crop_label.strip()
    pset = set(pred)

    if len(pset) == 0:
        return "MISS"
    if pset == {"unknown"} and gt != "unknown":
        return "UNKNOWN"
    if gt and (pset == {gt}):
        return "PERFECT"
    if gt and (gt in pset):
        return "HALLUCINATION" if len(pset - {gt}) > 0 else "PERFECT"
    return "HALLUCINATION"

def find_first_existing(paths: List[str]) -> Optional[str]:
    for p in paths:
        if p and os.path.exists(p):
            return p
    return None

def find_baseline_image(split: str, image_id: str) -> Optional[str]:
    return find_first_existing([
        f"datasets/{split}/images/{image_id}.jpg",
        f"datasets/{split}/images/{image_id}.jpeg",
        f"datasets/{split}/images/{image_id}.png",
    ])

def find_mask_rgb(split: str, image_id: str) -> Optional[str]:
    return find_first_existing([
        f"datasets/{split}/mask_rgb/{image_id}_mask.jpg",
        f"datasets/{split}/mask_rgb/{image_id}_mask.png",
        f"datasets/{split}/mask_rgb/{image_id}_mask.jpeg",
    ])

def list_crop_files(split: str, image_id: str) -> List[str]:
    d = f"datasets/crops_gt/{split}/{image_id}"
    if not os.path.isdir(d):
        return []
    return sorted([os.path.basename(p) for p in glob.glob(os.path.join(d, "*.png"))])

def path_crop_black(split: str, image_id: str, crop_file: str) -> Optional[str]:
    return find_first_existing([f"datasets/crops_gt/{split}/{image_id}/{crop_file}"])

def path_crop_white(split: str, image_id: str, crop_file: str) -> Optional[str]:
    return find_first_existing([f"datasets/crops_gt_white/{split}/{image_id}/{crop_file}"])

def load_per_crop_xlsx(xlsx_path: str) -> Dict[Tuple[str, str, str, str, str], Dict[str, Any]]:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    if "per_crop" not in wb.sheetnames:
        raise RuntimeError(f"Planilha sem aba 'per_crop'. Abas: {wb.sheetnames}")

    ws = wb["per_crop"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = {str(h).strip(): i for i, h in enumerate(header)}

    required = [
        "split","variant","image_id","crop_file","pred_labels","model",
        "fp_labels_not_in_gt","n_fp","gt_labels_image","crop_gt_label_from_filename"
    ]
    missing = [c for c in required if c not in idx]
    if missing:
        raise RuntimeError(f"Colunas ausentes em per_crop: {missing}")

    data: Dict[Tuple[str, str, str, str, str], Dict[str, Any]] = {}
    for r in rows:
        split = _to_str(r[idx["split"]]).strip()
        variant = _to_str(r[idx["variant"]]).strip()
        image_id = _to_str(r[idx["image_id"]]).strip()
        crop_file = _to_str(r[idx["crop_file"]]).strip()
        model = _to_str(r[idx["model"]]).strip()

        key = (split, variant, image_id, crop_file, model)
        data[key] = {
            "pred_labels": r[idx["pred_labels"]],
            "fp_labels_not_in_gt": r[idx["fp_labels_not_in_gt"]],
            "n_fp": r[idx["n_fp"]],
            "gt_labels_image": r[idx["gt_labels_image"]],
            "crop_gt_label_from_filename": r[idx["crop_gt_label_from_filename"]],
        }
    return data

def lookup_row(per_crop, split, variant_key, image_id, crop_file, model):
    for v in VARIANT_ALIASES.get(variant_key, [variant_key]):
        k = (split, v, image_id, crop_file, model)
        if k in per_crop:
            return per_crop[k]
    return {}

def draw_image_fit(c: canvas.Canvas, img_path: str, x: float, y: float, w: float, h: float) -> None:
    ir = ImageReader(img_path)
    iw, ih = ir.getSize()
    if iw <= 0 or ih <= 0:
        return
    scale = min(w / iw, h / ih)
    nw, nh = iw * scale, ih * scale
    ox = x + (w - nw) / 2.0
    oy = y + (h - nh) / 2.0
    c.drawImage(ir, ox, oy, width=nw, height=nh, preserveAspectRatio=True, mask='auto')

def write_wrapped(c: canvas.Canvas, text: str, x: float, y: float, max_chars: int, line_height: float) -> float:
    if not text:
        return y
    words = text.split(" ")
    line = ""
    for w in words:
        if len(line) + len(w) + 1 <= max_chars:
            line = (line + " " + w).strip()
        else:
            c.drawString(x, y, line)
            y -= line_height
            line = w
    if line:
        c.drawString(x, y, line)
        y -= line_height
    return y

def baseline_metrics(gt_set: Set[str], pred: List[str], n_fp: int):
    pred_set = set(pred)
    pred_n = len(pred)
    hit_any = (len(pred_set & gt_set) > 0) if gt_set else False
    hit_with_fp = (hit_any and n_fp > 0)
    fp_pct = (n_fp / pred_n) if pred_n > 0 else 0.0
    miss_n = len(gt_set - pred_set) if gt_set else 0
    perfect = (pred_set == gt_set) and bool(gt_set)
    return hit_any, hit_with_fp, fp_pct, pred_n, miss_n, perfect

def crop_metrics(gt_crop: str, pred: List[str], n_fp: int):
    pred_n = len(pred)
    hit = (gt_crop in pred)
    hit_with_fp = (hit and n_fp > 0)
    fp_pct = (n_fp / pred_n) if pred_n > 0 else 0.0
    status = compute_status_crop(gt_crop, pred)
    return hit, hit_with_fp, fp_pct, pred_n, status

def per_image_crop_stats(split, image_id, crop_files, per_crop, model, variant_key):
    total = 0
    hit = 0
    hit_fp = 0
    sum_fp_pct = 0.0
    sum_nfp = 0.0

    for crop_file in crop_files:
        gt = os.path.splitext(crop_file)[0]
        row = lookup_row(per_crop, split, variant_key, image_id, crop_file, model)
        if not row:
            continue
        pred = parse_labels(row.get("pred_labels"))
        nfp = int(row.get("n_fp") or 0)
        is_hit, is_hit_fp, fp_pct, _, _status = crop_metrics(gt, pred, nfp)

        total += 1
        if is_hit:
            hit += 1
            sum_fp_pct += fp_pct
            sum_nfp += nfp
            if is_hit_fp:
                hit_fp += 1

    cond = (hit_fp / hit) if hit > 0 else 0.0
    mean_fp_pct = (sum_fp_pct / hit) if hit > 0 else 0.0
    mean_nfp = (sum_nfp / hit) if hit > 0 else 0.0
    return total, hit, hit_fp, cond, mean_fp_pct, mean_nfp

def make_pdf_for_image(out_pdf, split, image_id, per_crop, models):
    os.makedirs(os.path.dirname(out_pdf), exist_ok=True)
    page_w, page_h = landscape(A4)
    c = canvas.Canvas(out_pdf, pagesize=(page_w, page_h))

    baseline_path = find_baseline_image(split, image_id)
    mask_path = find_mask_rgb(split, image_id)
    crop_files = list_crop_files(split, image_id)

    # GT image-level (da planilha)
    gt_image_str = ""
    gt_set = set()
    for m in models:
        row = lookup_row(per_crop, split, "baseline", image_id, "__full_image__", m)
        if row:
            gt_image_str = _to_str(row.get("gt_labels_image")).strip()
            gt_set = parse_gt_image_labels(gt_image_str)
            break

    def draw_header(title):
        c.setFont("Helvetica-Bold", 14)
        c.drawString(40, page_h - 40, title)
        c.setFont("Helvetica", 10)
        c.drawString(40, page_h - 60, f"split: {split}    image_id: {image_id}")
        if gt_image_str:
            c.drawString(40, page_h - 75, f"GT (image-level): {gt_image_str}")

    margin_x = 40
    top_y = page_h - 95
    img_box_w = (page_w - 2 * margin_x - 3 * 15) / 4.0
    img_box_h = 220
    gap = 15
    table_top = top_y - img_box_h - 20
    line_h = 13

    def draw_images_summary():
        # Baseline + Mask (sem bordas)
        labels = ["Baseline", "Mask"]
        paths = [baseline_path, mask_path]

        x = margin_x
        y = top_y - img_box_h
        box_w = (page_w - 2 * margin_x - gap) / 2.0

        for lab, p in zip(labels, paths):
            c.setFont("Helvetica-Bold", 10)
            c.drawString(x, top_y + 2, lab)

            if p and os.path.exists(p):
                draw_image_fit(c, p, x, y, box_w, img_box_h)
            else:
                c.setFont("Helvetica", 9)
                c.drawString(x + 4, y + img_box_h / 2, "(missing)")

            x += box_w + gap

    def draw_images_crops(crop_file: str):
        # Crop black + Crop white (sem bordas) — igual ao PDF antigo
        labels = ["Crop black", "Crop white"]
        paths = [
            path_crop_black(split, image_id, crop_file),
            path_crop_white(split, image_id, crop_file),
        ]

        x = margin_x
        y = top_y - img_box_h
        box_w = (page_w - 2 * margin_x - gap) / 2.0

        for lab, p in zip(labels, paths):
            c.setFont("Helvetica-Bold", 10)
            c.drawString(x, top_y + 2, lab)

            if p and os.path.exists(p):
                draw_image_fit(c, p, x, y, box_w, img_box_h)
            else:
                c.setFont("Helvetica", 9)
                c.drawString(x + 4, y + img_box_h / 2, "(missing)")

            x += box_w + gap

    # ===== Page 1: resumo + métricas =====
    draw_header("Per-image summary + conditional hallucination (when GT is present)")
    draw_images_summary()

    y = table_top
    c.setFont("Helvetica-Bold", 11)
    c.drawString(margin_x, y, "Baseline (image-level) + Crops (conditional hallucination on hits)")
    y -= 18
    c.setFont("Helvetica", 9)

    for model in models:
        # baseline row
        brow = lookup_row(per_crop, split, "baseline", image_id, "__full_image__", model)
        bpred = parse_labels(brow.get("pred_labels"))
        bnfp = int(brow.get("n_fp") or 0)
        hit_any, hit_fp, fp_pct, pred_n, miss_n, perfect = baseline_metrics(gt_set, bpred, bnfp)

        c.setFont("Helvetica-Bold", 9)
        c.drawString(margin_x, y, f"- {model}")
        y -= line_h
        c.setFont("Helvetica", 9)
        c.drawString(margin_x + 14, y,
                     f"baseline: hit_any={'YES' if hit_any else 'NO'} | perfect={'YES' if perfect else 'NO'} | miss_n={miss_n} | n_fp={bnfp} | pred_n={pred_n} | FP%={fp_pct*100:.1f}% | hit+FP={'YES' if hit_fp else 'NO'}")
        y -= line_h

        # crops black stats (por modelo)
        tb, hb, hfpb, condb, meanfpb, meannfpb = per_image_crop_stats(split, image_id, crop_files, per_crop, model, "crops_black")
        c.drawString(margin_x + 14, y,
                     f"crops_black: hits={hb}/{tb} | hit+FP={hfpb} | P(FP>0|hit)={condb*100:.1f}% | mean FP% on hit={meanfpb*100:.1f}% | mean n_fp on hit={meannfpb:.2f}")
        y -= line_h

        # crops white stats (por modelo)
        tw, hw, hfpw, condw, meanfpw, meannfpw = per_image_crop_stats(split, image_id, crop_files, per_crop, model, "crops_white")
        c.drawString(margin_x + 14, y,
                     f"crops_white: hits={hw}/{tw} | hit+FP={hfpw} | P(FP>0|hit)={condw*100:.1f}% | mean FP% on hit={meanfpw*100:.1f}% | mean n_fp on hit={meannfpw:.2f}")
        y -= (line_h + 6)

        if y < 60:
            c.showPage()
            draw_header("(cont.) Per-image summary")
            draw_images_summary()
            y = table_top
            c.setFont("Helvetica", 9)

    c.showPage()

    # ===== Pages: 1 por crop =====
    for i, crop_file in enumerate(crop_files, start=1):
        gt_crop = os.path.splitext(crop_file)[0]
        draw_header(f"Crop {i}/{len(crop_files)}: {crop_file} (GT crop-level: {gt_crop})")
        draw_images_crops(crop_file)

        y = table_top
        c.setFont("Helvetica-Bold", 11)
        c.drawString(margin_x, y, "Per-crop predictions (NEW: HIT and FP%)")
        y -= 18
        c.setFont("Helvetica", 9)

        for model in models:
            # baseline (sempre comparado ao GT image-level)
            #brow = lookup_row(per_crop, split, "baseline", image_id, "__full_image__", model)
            #bpred = parse_labels(brow.get("pred_labels"))
            #bnfp = int(brow.get("n_fp") or 0)
            #hit_any, hit_fp, b_fp_pct, b_pred_n, miss_n, perfect = baseline_metrics(gt_set, bpred, bnfp)

            # crop black / white (comparado ao GT do crop)
            blkrow = lookup_row(per_crop, split, "crops_black", image_id, crop_file, model)
            whtrow = lookup_row(per_crop, split, "crops_white", image_id, crop_file, model)

            kpred = parse_labels(blkrow.get("pred_labels"))
            knfp = int(blkrow.get("n_fp") or 0)
            whpred = parse_labels(whtrow.get("pred_labels"))
            whnfp = int(whtrow.get("n_fp") or 0)

            k_hit, k_hit_fp, k_fp_pct, k_pred_n, k_status = crop_metrics(gt_crop, kpred, knfp)
            w_hit, w_hit_fp, w_fp_pct, w_pred_n, w_status = crop_metrics(gt_crop, whpred, whnfp)

            c.setFont("Helvetica-Bold", 9)
            c.drawString(margin_x, y, f"- {model}")
            y -= line_h
            c.setFont("Helvetica", 9)

            y = write_wrapped(
                c,
                f"crop_black: {k_status} | hit={'YES' if k_hit else 'NO'} | n_fp={knfp} | pred_n={k_pred_n} | FP%={k_fp_pct*100:.1f}% | hit+FP={'YES' if k_hit_fp else 'NO'} | pred=[{','.join(kpred)}]",
                margin_x + 14, y, 190, line_h
            )
            y = write_wrapped(
                c,
                f"crop_white: {w_status} | hit={'YES' if w_hit else 'NO'} | n_fp={whnfp} | pred_n={w_pred_n} | FP%={w_fp_pct*100:.1f}% | hit+FP={'YES' if w_hit_fp else 'NO'} | pred=[{','.join(whpred)}]",
                margin_x + 14, y, 190, line_h
            )
            y -= 8

            if y < 60:
                c.showPage()
                draw_header(f"(cont.) Crop {i}/{len(crop_files)}: {crop_file}")
                draw_images_crops(crop_file)
                y = table_top
                c.setFont("Helvetica", 9)

        c.showPage()

    c.save()

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--splits", required=True, help="test_open,test_closed")
    ap.add_argument("--image_ids", default="", help="lista separada por vírgula; se vazio, gera para todas do split")
    ap.add_argument("--outdir", required=True)
    ap.add_argument("--tables_xlsx", default="DL_all_vlms_baseline_black_white_tables.xlsx")
    ap.add_argument("--models", default="gemma3,llama3_2_vision,qwen2_5_vl")
    args = ap.parse_args()

    splits = [s.strip() for s in args.splits.split(",") if s.strip()]
    models = [m.strip() for m in args.models.split(",") if m.strip()]

    if not os.path.exists(args.tables_xlsx):
        raise SystemExit(f"[ERRO] Não achei a planilha: {args.tables_xlsx}")

    per_crop = load_per_crop_xlsx(args.tables_xlsx)

    for split in splits:
        if args.image_ids.strip():
            image_ids = [x.strip() for x in args.image_ids.split(",") if x.strip()]
        else:
            img_dir = f"datasets/{split}/images"
            image_ids = sorted([os.path.splitext(p)[0] for p in os.listdir(img_dir) if p.lower().endswith(".jpg")])

        for image_id in image_ids:
            out_pdf = os.path.join(args.outdir, split, f"{image_id}.pdf")
            try:
                make_pdf_for_image(out_pdf, split, image_id, per_crop, models)
                print(f"[OK] {out_pdf}")
            except Exception as e:
                print(f"[ERRO] {split}/{image_id}: {e}")

if __name__ == "__main__":
    main()
